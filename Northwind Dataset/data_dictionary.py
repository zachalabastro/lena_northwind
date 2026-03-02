#!/usr/bin/env python3
"""
data_dictionary.py
==================
Generate a styled Excel data dictionary from any pandas DataFrame.

Workflow
--------
  DataFrame  ──►  ydata_profiling (JSON)  ──►  parsed stats  ──►  styled .xlsx

Usage — as a script
-------------------
  # From a CSV file
  python data_dictionary.py --csv mydata.csv --output my_dict.xlsx

  # From a SQLite database
  python data_dictionary.py --sqlite northwind.db --query "SELECT * FROM Orders" --output orders_dict.xlsx

  # Full Northwind enriched dataset (reproduces the notebook exactly)
  python data_dictionary.py --northwind northwind.db --output northwind_data_dictionary.xlsx

Usage — as a library
--------------------
  from data_dictionary import build_data_dictionary

  df = pd.read_csv("mydata.csv")
  build_data_dictionary(df, "my_dict.xlsx", title="My Dataset")
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from ydata_profiling import ProfileReport


# ── Column-group definitions ──────────────────────────────────────────────────
# Each tuple: (group label, [column names]).
# Columns absent from the data dictionary are silently skipped.
COLUMN_GROUPS: list[tuple[str, list[str]]] = [
    ("Identity",          ["Type", "Pandas dtype", "Memory (bytes)"]),
    ("Completeness",      ["Total Rows", "Count (non-null)", "# Missing", "% Missing"]),
    ("Uniqueness",        ["# Distinct", "% Distinct", "# Unique (exact)", "% Unique (exact)", "Is Unique"]),
    ("Central Tendency",  ["Mean", "Median (50%)"]),
    ("Spread",            ["Std Dev", "Variance", "MAD", "CV", "IQR"]),
    ("Shape",             ["Skewness", "Kurtosis"]),
    ("Range / Quantiles", ["Min", "5th Pct", "25th Pct", "75th Pct", "95th Pct", "Max", "Range"]),
    ("Aggregate",         ["Sum"]),
    ("Special Counts",    ["# Zeros", "% Zeros", "# Negative", "% Negative", "# Infinite", "Monotonic"]),
    ("Text Metrics",      ["Min Length", "Max Length", "Mean Length", "Median Length",
                           "# Characters", "# Distinct Chars"]),
    ("Most Frequent",     ["Top Value", "Top Freq"]),
]

# Columns that receive the per-column blue intensity gradient
GRADIENT_COLS: list[str] = [
    "Mean", "Median (50%)", "Std Dev", "Variance", "MAD",
    "CV", "IQR", "Skewness", "Kurtosis", "Sum", "Range",
]

# Excel number formats applied per column
NUM_FMT: dict[str, str] = {
    "% Missing":      '0.00"%"', "% Distinct":     '0.00"%"',
    "% Unique (exact)":'0.00"%"',"% Zeros":         '0.00"%"',
    "% Negative":     '0.00"%"',
    "Mean":           "#,##0.0000", "Median (50%)":  "#,##0.0000",
    "Std Dev":        "#,##0.0000", "Variance":      "#,##0.0000",
    "MAD":            "#,##0.0000", "CV":            "#,##0.0000",
    "IQR":            "#,##0.0000", "Skewness":      "#,##0.0000",
    "Kurtosis":       "#,##0.0000", "5th Pct":       "#,##0.0000",
    "25th Pct":       "#,##0.0000", "75th Pct":      "#,##0.0000",
    "95th Pct":       "#,##0.0000", "Range":         "#,##0.0000",
    "Sum":            "#,##0.00",   "Mean Length":   "0.00",
    "Median Length":  "0.00",       "Total Rows":    "#,##0",
    "Count (non-null)":"#,##0",     "# Missing":     "#,##0",
    "# Distinct":     "#,##0",      "# Unique (exact)":"#,##0",
    "# Zeros":        "#,##0",      "# Negative":    "#,##0",
    "# Infinite":     "#,##0",      "Memory (bytes)":"#,##0",
    "# Characters":   "#,##0",      "# Distinct Chars":"#,##0",
    "Top Freq":       "#,##0",
}

# Excel fill colours (hex, no leading #)
PALETTE = dict(
    hdr_bg="0f172a", hdr_fg="cbd5e1", grp_fg="93c5fd",  # header / group labels
    idx_bg="1e293b", idx_fg="e2e8f0",                    # Variable column
    odd_bg="f8fafc", even_bg="ffffff",                   # alternating data rows
    num_bg="1d4ed8", txt_bg="0f766e", dt_bg="7c3aed",   # type badge fills
    bool_bg="b45309", cat_bg="15803d", unk_bg="64748b",
)
GROUP_BG = [
    "1e3a5f", "164e3f", "3b1f63", "4a1c1c", "1c3a4a",
    "2a2a10", "1a3a2a", "1f2a40", "2f1f40", "1a3030", "10302a",
]
TYPE_BADGE_FILL = {
    "Numeric":     PALETTE["num_bg"],
    "Text":        PALETTE["txt_bg"],
    "DateTime":    PALETTE["dt_bg"],
    "Boolean":     PALETTE["bool_bg"],
    "Categorical": PALETTE["cat_bg"],
}


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — Profile the DataFrame and extract per-variable stats into a table
# ─────────────────────────────────────────────────────────────────────────────

def _get(d: dict, key: str, default=np.nan):
    """Safe dict lookup — returns *default* when key is absent or None."""
    v = d.get(key)
    return default if v is None else v


def _top(vc: dict) -> tuple:
    """Return (most_frequent_value, count) from a value_counts dict."""
    if isinstance(vc, dict) and vc:
        k, v = next(iter(vc.items()))
        return k, int(v)
    return np.nan, np.nan


def _pct(v) -> float:
    """Convert a [0, 1] fraction to a percentage rounded to 4 dp."""
    return round(float(v) * 100, 4) if not (isinstance(v, float) and np.isnan(v)) else np.nan


def _r(v, dp: int = 4) -> float:
    """Round *v* to *dp* decimal places; return NaN on failure."""
    try:
        f = float(v)
        return round(f, dp) if not np.isnan(f) else np.nan
    except Exception:
        return np.nan


def _safe_int(s: dict, key: str):
    """Extract an integer stat from a variable's stat dict."""
    v = _get(s, key)
    return int(v) if not isinstance(v, float) else np.nan


def profile_to_dict_df(df: pd.DataFrame, minimal: bool = False) -> pd.DataFrame:
    """
    Run ydata_profiling on *df* and return a tidy data dictionary DataFrame.

    Parameters
    ----------
    df : pd.DataFrame
        The dataset to profile.
    minimal : bool
        Pass ``True`` to skip expensive correlations / interactions
        (much faster, same per-variable stats).

    Returns
    -------
    pd.DataFrame
        Index = variable name.  Columns = all statistical attributes.
        Cells that are N/A for a given variable type are ``NaN``.
    """
    profile = ProfileReport(df, minimal=minimal, progress_bar=True)
    variables: dict = json.loads(profile.to_json()).get("variables", {})

    rows = []
    for col_name, s in variables.items():
        vtype     = _get(s, "type", "Unknown")
        n         = _get(s, "n")
        n_missing = _get(s, "n_missing")
        n_count   = (
            int(n) - int(n_missing)
            if not isinstance(n, float) and not isinstance(n_missing, float)
            else np.nan
        )
        top_val, top_freq = _top(_get(s, "value_counts_without_nan", {}))

        rows.append({
            # ── identity ──────────────────────────────────────────────────
            "Variable":         col_name,
            "Type":             vtype,
            "Pandas dtype":     str(df[col_name].dtype) if col_name in df.columns else "—",
            "Memory (bytes)":   _safe_int(s, "memory_size"),
            # ── completeness ──────────────────────────────────────────────
            "Total Rows":       int(n) if not isinstance(n, float) else np.nan,
            "Count (non-null)": int(n_count) if not isinstance(n_count, float) else n_count,
            "# Missing":        _safe_int(s, "n_missing"),
            "% Missing":        _pct(_get(s, "p_missing")),
            # ── uniqueness ────────────────────────────────────────────────
            "# Distinct":       _safe_int(s, "n_distinct"),
            "% Distinct":       _pct(_get(s, "p_distinct")),
            "# Unique (exact)": _safe_int(s, "n_unique"),
            "% Unique (exact)": _pct(_get(s, "p_unique")),
            "Is Unique":        bool(_get(s, "is_unique")) if not isinstance(_get(s, "is_unique"), float) else np.nan,
            # ── central tendency (Numeric) ────────────────────────────────
            "Mean":             _r(_get(s, "mean")),
            "Median (50%)":     _r(_get(s, "50%")),
            # ── spread (Numeric) ─────────────────────────────────────────
            "Std Dev":          _r(_get(s, "std")),
            "Variance":         _r(_get(s, "variance")),
            "MAD":              _r(_get(s, "mad")),
            "CV":               _r(_get(s, "cv")),
            "IQR":              _r(_get(s, "iqr")),
            # ── shape (Numeric) ──────────────────────────────────────────
            "Skewness":         _r(_get(s, "skewness")),
            "Kurtosis":         _r(_get(s, "kurtosis")),
            # ── range / quantiles (Numeric) ───────────────────────────────
            "Min":              _get(s, "min"),
            "5th Pct":          _r(_get(s, "5%")),
            "25th Pct":         _r(_get(s, "25%")),
            "75th Pct":         _r(_get(s, "75%")),
            "95th Pct":         _r(_get(s, "95%")),
            "Max":              _get(s, "max"),
            "Range":            _r(_get(s, "range")),
            # ── aggregate (Numeric) ───────────────────────────────────────
            "Sum":              _r(_get(s, "sum"), dp=2),
            # ── special counts (Numeric) ──────────────────────────────────
            "# Zeros":          _safe_int(s, "n_zeros"),
            "% Zeros":          _pct(_get(s, "p_zeros")),
            "# Negative":       _safe_int(s, "n_negative"),
            "% Negative":       _pct(_get(s, "p_negative")),
            "# Infinite":       _safe_int(s, "n_infinite"),
            "Monotonic":        _get(s, "monotonic"),
            # ── text-specific ─────────────────────────────────────────────
            "Min Length":       _safe_int(s, "min_length"),
            "Max Length":       _safe_int(s, "max_length"),
            "Mean Length":      _r(_get(s, "mean_length")),
            "Median Length":    _r(_get(s, "median_length")),
            "# Characters":     _safe_int(s, "n_characters"),
            "# Distinct Chars": _safe_int(s, "n_characters_distinct"),
            # ── most frequent value ───────────────────────────────────────
            "Top Value":        top_val,
            "Top Freq":         int(top_freq) if not (isinstance(top_freq, float) and np.isnan(top_freq)) else np.nan,
        })

    return pd.DataFrame(rows).set_index("Variable")


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — Write the styled Excel workbook
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(hex_color: str = "1a1a2e", bold: bool = False, sz: int = 10) -> Font:
    return Font(name="Calibri", color=hex_color, bold=bold, size=sz)


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border() -> Border:
    side = Side(style="thin", color="d0d7de")
    return Border(left=side, right=side, top=side, bottom=side)


def _lerp_color(t: float,
                lo: tuple = (239, 246, 255),
                hi: tuple = (29, 78, 216)) -> str:
    """Interpolate between two RGB colours; return an openpyxl hex string."""
    r = int(lo[0] + (hi[0] - lo[0]) * t)
    g = int(lo[1] + (hi[1] - lo[1]) * t)
    b = int(lo[2] + (hi[2] - lo[2]) * t)
    return f"{r:02X}{g:02X}{b:02X}"


def _missing_color(pct: float) -> str:
    """White → red gradient for the % Missing column."""
    t = min(float(pct) / 100, 1.0) * 0.85
    g = int(255 * (1 - t))
    b = int(255 * (1 - t))
    return f"FF{g:02X}{b:02X}"


def build_excel(dict_df: pd.DataFrame, out_path: str | Path) -> None:
    """
    Write *dict_df* (output of :func:`profile_to_dict_df`) to a styled Excel file.

    Layout
    ------
    Row 1 : Column-group labels (merged cells, dark tinted background per group).
    Row 2 : Column names (deep navy, white text, word-wrapped).
    Row 3+ : One data row per variable.
    Col A : Variable names (dark slate, bold, sticky / frozen).

    Styling highlights
    ------------------
    * Type badge  — solid fill per profiling type (Numeric=blue, Text=teal, …)
    * % Missing   — white → red heat map
    * Spread cols — per-column light-blue → dark-blue intensity gradient
    * Alternating — pale-blue / white data rows
    * Freeze      — pane locked at B3 (headers + Variable column always visible)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Dictionary"
    ws.sheet_properties.tabColor = "1d4ed8"

    all_cols = list(dict_df.columns)

    # ── Row 1 : Group headers ─────────────────────────────────────────────────
    # Column A spans rows 1–2 (the "VARIABLE" label)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    vc = ws.cell(row=1, column=1, value="VARIABLE")
    vc.fill      = _fill(PALETTE["hdr_bg"])
    vc.font      = _font(PALETTE["grp_fg"], bold=True, sz=9)
    vc.alignment = _align("center", "center")
    vc.border    = _border()

    col_offset = 2
    for gi, (grp_name, grp_cols) in enumerate(COLUMN_GROUPS):
        present = [c for c in grp_cols if c in dict_df.columns]
        if not present:
            continue
        s, e = col_offset, col_offset + len(present) - 1
        if s < e:
            ws.merge_cells(start_row=1, start_column=s, end_row=1, end_column=e)
        # Only the top-left cell of a merged range accepts values/styles
        gc = ws.cell(row=1, column=s, value=grp_name.upper())
        gc.fill      = _fill(GROUP_BG[gi % len(GROUP_BG)])
        gc.font      = _font(PALETTE["grp_fg"], bold=True, sz=9)
        gc.alignment = _align("center", "center")
        gc.border    = _border()
        col_offset   = e + 1

    # ── Row 2 : Column headers ────────────────────────────────────────────────
    for ci, col_name in enumerate(all_cols):
        cell = ws.cell(row=2, column=ci + 2, value=col_name)
        cell.fill      = _fill(PALETTE["hdr_bg"])
        cell.font      = _font(PALETTE["hdr_fg"], bold=True, sz=9)
        cell.alignment = _align("center", "center", wrap=True)
        cell.border    = _border()

    # ── Pre-compute gradient value ranges ─────────────────────────────────────
    col_ranges: dict[str, tuple[float, float]] = {}
    for c in GRADIENT_COLS:
        if c in dict_df.columns:
            nums = pd.to_numeric(dict_df[c], errors="coerce").dropna()
            if not nums.empty and nums.max() != nums.min():
                col_ranges[c] = (nums.min(), nums.max())

    missing_ci = (all_cols.index("% Missing") + 2) if "% Missing" in all_cols else None
    type_ci    = (all_cols.index("Type") + 2)       if "Type"      in all_cols else None

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, (var_name, row_data) in enumerate(dict_df.iterrows()):
        er     = ri + 3
        row_bg = PALETTE["odd_bg"] if ri % 2 == 0 else PALETTE["even_bg"]

        # Variable name cell (column A)
        idx            = ws.cell(row=er, column=1, value=var_name)
        idx.fill       = _fill(PALETTE["idx_bg"])
        idx.font       = _font(PALETTE["idx_fg"], bold=True, sz=10)
        idx.alignment  = _align("left", "center")
        idx.border     = _border()

        for ci, col_name in enumerate(all_cols):
            ec  = ci + 2
            raw = row_data[col_name]
            # Coerce NaN / numpy scalars to Python-native types for openpyxl
            val = (
                None
                if (pd.isna(raw) if not isinstance(raw, (list, dict)) else False)
                else (raw.item() if hasattr(raw, "item") else raw)
            )

            cell           = ws.cell(row=er, column=ec, value=val)
            cell.alignment = _align("center", "center")
            cell.border    = _border()
            cell.font      = _font("1a1a2e", sz=10)

            if col_name in NUM_FMT and val is not None:
                cell.number_format = NUM_FMT[col_name]

            # Priority: type badge > missing heat > gradient > default
            if ec == type_ci and isinstance(val, str):
                cell.fill = _fill(TYPE_BADGE_FILL.get(val, PALETTE["unk_bg"]))
                cell.font = _font("ffffff", bold=True, sz=10)
            elif ec == missing_ci and val is not None:
                cell.fill = _fill(_missing_color(val))
                cell.font = _font("ffffff" if val > 55 else "1a1a1a", sz=10)
            elif col_name in col_ranges and val is not None:
                mn, mx    = col_ranges[col_name]
                t         = (float(val) - mn) / (mx - mn)
                cell.fill = _fill(_lerp_color(t))
                cell.font = _font("ffffff" if t > 0.55 else "1a1a2e", sz=10)
            else:
                cell.fill = _fill(row_bg)

    # ── Freeze panes, dimensions ──────────────────────────────────────────────
    ws.freeze_panes = "B3"
    ws.row_dimensions[1].height = 20   # group header row
    ws.row_dimensions[2].height = 34   # column header row (allows wrap)
    for r in range(3, 3 + len(dict_df)):
        ws.row_dimensions[r].height = 17

    ws.column_dimensions["A"].width = 22
    for ci, col_name in enumerate(all_cols):
        letter  = get_column_letter(ci + 2)
        lengths = [len(str(col_name))] + [len(str(v)) for v in dict_df[col_name].dropna()]
        ws.column_dimensions[letter].width = min(max(max(lengths) * 1.05 + 2, 9), 26)

    wb.save(out_path)
    print(f"  Saved → {out_path}")
    print(f"  {len(dict_df)} variables × {len(all_cols)} attributes")
    print(f"  Freeze panes at B3  |  Sheet tab: blue")


# ─────────────────────────────────────────────────────────────────────────────
# Public convenience wrapper
# ─────────────────────────────────────────────────────────────────────────────

def build_data_dictionary(
    df: pd.DataFrame,
    out_path: str | Path = "data_dictionary.xlsx",
    title: str = "Data Dictionary",
    minimal: bool = False,
) -> pd.DataFrame:
    """
    Profile *df* and write a styled Excel data dictionary.

    Parameters
    ----------
    df       : The dataset to document.
    out_path : Destination ``.xlsx`` file path.
    title    : Used as the ydata_profiling report title (cosmetic).
    minimal  : ``True`` skips expensive correlation steps (faster, same stats).

    Returns
    -------
    pd.DataFrame
        The raw data dictionary table (useful for further processing).

    Example
    -------
    >>> import pandas as pd
    >>> from data_dictionary import build_data_dictionary
    >>> df = pd.read_csv("sales.csv")
    >>> dict_df = build_data_dictionary(df, "sales_dict.xlsx")
    """
    print(f"Profiling {df.shape[0]:,} rows × {df.shape[1]} columns …")
    dict_df = profile_to_dict_df(df, minimal=minimal)
    print(f"Building Excel workbook …")
    build_excel(dict_df, out_path)
    return dict_df


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _build_northwind_df(db_path: str) -> pd.DataFrame:
    """Reproduce the enriched line-item DataFrame from northwind_data_dictionary.ipynb."""
    conn             = sqlite3.connect(db_path)
    orders_df        = pd.read_sql_query("SELECT * FROM Orders;", conn)
    order_details_df = pd.read_sql_query('SELECT * FROM "Order Details";', conn)
    products_df      = pd.read_sql_query("SELECT * FROM Products;", conn)
    customers_df     = pd.read_sql_query("SELECT * FROM Customers;", conn)
    conn.close()

    enriched = (
        orders_df
        .merge(order_details_df, on="OrderID",    how="left")
        .merge(products_df,      on="ProductID",  how="left")
        .merge(customers_df,     on="CustomerID", how="left")
    )
    df = (
        enriched
        .drop(columns=[c for c in enriched.columns if c.endswith("_y")])
        .rename(columns=lambda x: x.replace("_x", ""))
    )
    df["OrderRevenue"] = (
        (df["UnitPrice"] - df["UnitPrice"] * df["Discount"]) * df["Quantity"]
    )
    return df.reset_index(names="LineItem")


def main(argv=None):
    parser = argparse.ArgumentParser(
        prog="data_dictionary",
        description="Generate a styled Excel data dictionary from a DataFrame.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python data_dictionary.py --csv sales.csv
  python data_dictionary.py --csv sales.csv --output sales_dict.xlsx --minimal
  python data_dictionary.py --sqlite mydb.db --query "SELECT * FROM orders"
  python data_dictionary.py --northwind northwind.db
        """,
    )

    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--csv",       metavar="FILE",   help="Path to a CSV file.")
    src.add_argument("--sqlite",    metavar="FILE",   help="Path to a SQLite database file.")
    src.add_argument("--northwind", metavar="FILE",   help="Reproduce the Northwind enriched dataset.")

    parser.add_argument("--query",   metavar="SQL",    help="SQL query (required with --sqlite).",
                        default="")
    parser.add_argument("--output",  metavar="FILE",   default="data_dictionary.xlsx",
                        help="Output Excel path (default: data_dictionary.xlsx).")
    parser.add_argument("--minimal", action="store_true",
                        help="Skip expensive correlations for faster profiling.")
    parser.add_argument("--title",   metavar="TEXT",   default="Data Dictionary",
                        help="Report title (cosmetic).")

    args = parser.parse_args(argv)

    # ── Load data ─────────────────────────────────────────────────────────────
    if args.csv:
        print(f"Loading CSV: {args.csv}")
        df = pd.read_csv(args.csv)

    elif args.sqlite:
        if not args.query:
            parser.error("--sqlite requires --query 'SELECT …'")
        print(f"Loading SQLite: {args.sqlite}  query: {args.query!r}")
        conn = sqlite3.connect(args.sqlite)
        df   = pd.read_sql_query(args.query, conn)
        conn.close()

    else:  # --northwind
        print(f"Building Northwind enriched dataset from: {args.northwind}")
        df = _build_northwind_df(args.northwind)

    print(f"Shape: {df.shape[0]:,} rows × {df.shape[1]} columns")
    build_data_dictionary(df, out_path=args.output, title=args.title, minimal=args.minimal)


if __name__ == "__main__":
    main()
